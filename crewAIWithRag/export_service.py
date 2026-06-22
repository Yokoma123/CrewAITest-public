import os
from datetime import datetime
from typing import Dict, List

import student_store


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.getenv("STUDENT_INFO_DATA_DIR", os.path.join(BASE_DIR, "data"))
EXPORT_DIR = os.path.join(DATA_DIR, "exports")


def export_excel(include: List[str], filters: Dict[str, str], selected_fields: List[str] | None = None) -> str:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ValueError("导出 Excel 需要安装 openpyxl") from exc

    os.makedirs(EXPORT_DIR, exist_ok=True)
    include = include or ["students"]
    students = student_store.list_students(filters.get("query", ""), filters)
    dynamic_fields = student_store.get_dynamic_fields()

    wb = Workbook()
    ws = wb.active
    ws.title = "学生信息"
    student_headers = [
        ("name", "姓名"),
        ("class_name", "班级"),
        ("student_id", "学号"),
        ("gender", "性别"),
        ("phone", "联系电话"),
        ("dorm_location", "寝室位置"),
        ("dorm_room", "寝室号"),
        ("advisor_name", "导师姓名"),
        ("political_status", "政治面貌"),
        ("tripartite_status", "是否签署三方"),
        ("destination_type", "去向类型"),
        ("employer_name", "单位名称"),
        ("job_title", "岗位"),
        ("job_city", "城市"),
        ("is_further_study", "是否升学"),
        ("destination_note", "去向备注"),
    ]
    if selected_fields:
        label_by_key = dict(student_headers)
        student_headers = [(key, label_by_key[key]) for key in selected_fields if key in label_by_key]
    extra_headers = [(field["field_key"], field["label"]) for field in dynamic_fields]
    if selected_fields:
        allowed = set(selected_fields)
        extra_headers = [(key, label) for key, label in extra_headers if key in allowed]
    ws.append([label for _, label in student_headers + extra_headers])
    for student in students:
        extra = student.get("extra_values", {})
        ws.append(
            [student.get(key, "") for key, _ in student_headers]
            + [extra.get(key, "") for key, _ in extra_headers]
        )

    if "records" in include:
        record_ws = wb.create_sheet("奖惩记录")
        record_ws.append(["学生姓名", "学号", "类型", "标题", "日期", "说明", "记录人/来源"])
        for student in students:
            for record in student.get("records", []):
                record_ws.append(
                    [
                        student.get("name", ""),
                        student.get("student_id", ""),
                        record.get("record_type", ""),
                        record.get("title", ""),
                        record.get("record_date", ""),
                        record.get("description", ""),
                        record.get("source", ""),
                    ]
                )

    if "activities" in include:
        activity_ws = wb.create_sheet("日常活动")
        activity_ws.append(["学生姓名", "学号", "时间", "任务名称", "任务量", "时长", "分数", "计算规则", "备注"])
        for student in students:
            for activity in student.get("activities", []):
                activity_ws.append(
                    [
                        student.get("name", ""),
                        student.get("student_id", ""),
                        activity.get("activity_date", ""),
                        activity.get("task_name", ""),
                        activity.get("task_quantity", ""),
                        activity.get("duration_hours", ""),
                        activity.get("score", ""),
                        activity.get("score_rule", ""),
                        activity.get("note", ""),
                    ]
                )

    for sheet in wb.worksheets:
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 32)
            sheet.column_dimensions[column[0].column_letter].width = width

    filename = f"学生信息导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join(EXPORT_DIR, filename)
    wb.save(path)
    return path
