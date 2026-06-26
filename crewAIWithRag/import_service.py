import csv
import os
import re
from typing import Any, Dict, List, Tuple

import student_store


FIELD_ALIASES = {
    "name": ["姓名", "学生姓名", "名字", "name"],
    "class_name": ["班级", "所在班级", "行政班", "class", "class_name"],
    "student_id": ["学号", "学号（职工号）", "学号(职工号)", "学生学号", "职工号", "student_id", "studentid", "id"],
    "phone": ["联系电话", "电话", "移动电话", "手机号", "手机号码", "联系方式", "phone"],
    "dorm_location": ["寝室位置", "宿舍位置", "所在寝室楼", "所在宿舍楼", "寝室楼", "宿舍楼", "公寓", "dorm_location"],
    "dorm_room": ["寝室号", "宿舍号", "房间号", "寝室", "宿舍", "dorm_room"],
    "advisor_name": ["导师姓名", "导师", "指导老师", "辅导员", "班主任", "advisor_name"],
    "gender": ["性别", "gender"],
    "political_status": ["政治面貌", "政治身份", "political_status"],
    "tripartite_status": ["是否签署三方", "三方", "三方协议", "签署三方", "tripartite_status"],
    "destination_type": ["去向类型", "毕业去向", "就业去向", "去向", "destination_type"],
    "employer_name": ["单位名称", "就业单位", "签约单位", "工作单位", "employer_name"],
    "job_title": ["岗位", "职位", "工作岗位", "job_title"],
    "job_city": ["城市", "工作城市", "就业城市", "job_city"],
    "is_further_study": ["是否升学", "升学", "is_further_study"],
    "destination_note": ["去向备注", "就业备注", "备注", "destination_note"],
}


IGNORED_HEADERS = {"序号", "编号", "证件号码", "身份证号"}


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    return re.sub(r"[\s_()（）【】\[\]{}<>《》:：,，、/\\\-]+", "", text)


def header_matches(candidate: str, aliases: List[str]) -> bool:
    normalized_candidate = normalize_header(candidate)
    if not normalized_candidate:
        return False
    for alias in aliases:
        normalized_alias = normalize_header(alias)
        if not normalized_alias:
            continue
        if normalized_candidate == normalized_alias:
            return True
        if len(normalized_alias) >= 3 and normalized_alias in normalized_candidate:
            return True
        if len(normalized_candidate) >= 3 and normalized_candidate in normalized_alias:
            return True
    return False


def header_exact_matches(candidate: str, aliases: List[str]) -> bool:
    normalized_candidate = normalize_header(candidate)
    return any(normalized_candidate == normalize_header(alias) for alias in aliases if normalize_header(alias))


def find_matching_header(headers: List[str], aliases: List[str], used_headers: set, exact_only: bool = False) -> str:
    for header in headers:
        if header in used_headers:
            continue
        if header_exact_matches(header, aliases) if exact_only else header_matches(header, aliases):
            return header
    return ""


def make_dynamic_key(header: str) -> str:
    cleaned = re.sub(r"\W+", "_", header.strip().lower(), flags=re.UNICODE).strip("_")
    return f"extra_{cleaned or 'field'}"


def build_header_map(headers: List[str]) -> Tuple[Dict[str, str], Dict[str, str]]:
    header_map = {}
    used_headers = set()
    field_aliases = [
        (field, aliases + [student_store.FIELD_LABELS.get(field, field)])
        for field, aliases in FIELD_ALIASES.items()
    ]
    for field, aliases in field_aliases:
        original = find_matching_header(headers, aliases, used_headers, exact_only=True)
        if original:
            header_map[field] = original
            used_headers.add(original)
    for field, aliases in field_aliases:
        if field in header_map:
            continue
        original = find_matching_header(headers, aliases, used_headers)
        if original:
            header_map[field] = original
            used_headers.add(original)

    dynamic_headers = {}
    for header in headers:
        header = str(header or "").strip()
        if not header or header in used_headers or header in IGNORED_HEADERS:
            continue
        dynamic_headers[make_dynamic_key(header)] = header
    return header_map, dynamic_headers


def read_csv(path: str) -> List[Dict[str, Any]]:
    encodings = ["utf-8-sig", "utf-8", "gbk"]
    last_error = None
    for encoding in encodings:
        try:
            with open(path, "r", encoding=encoding, newline="") as file:
                return list(csv.DictReader(file))
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"无法读取 CSV 文件编码: {last_error}")


def read_excel(path: str) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("读取 Excel 需要安装 openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    result = []
    for row in rows[1:]:
        item = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            item[header] = "" if index >= len(row) or row[index] is None else str(row[index]).strip()
        result.append(item)
    return result


def read_xls(path: str) -> List[Dict[str, Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("读取老版 .xls 文件需要安装 xlrd，或请先另存为 .xlsx 后再导入") from exc

    try:
        workbook = xlrd.open_workbook(path)
    except Exception as exc:
        raise ValueError(f"无法读取 .xls 文件，请确认文件未损坏: {exc}") from exc

    if not workbook.nsheets:
        return []
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        return []
    headers = [str(sheet.cell_value(0, col) or "").strip() for col in range(sheet.ncols)]
    result = []
    for row_index in range(1, sheet.nrows):
        item = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            cell = sheet.cell(row_index, col_index)
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_NUMBER and float(value).is_integer():
                value = str(int(value))
            else:
                value = "" if value is None else str(value).strip()
            item[header] = value
        result.append(item)
    return result


def read_rows(path: str) -> List[Dict[str, Any]]:
    suffix = os.path.splitext(path)[1].lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix == ".xlsx":
        return read_excel(path)
    if suffix == ".xls":
        return read_xls(path)
    raise ValueError("暂只支持 CSV、XLSX、XLS 文件")


def import_students(path: str, filename: str = "") -> Dict[str, Any]:
    if not os.path.exists(path):
        raise ValueError("导入文件不存在")
    rows = read_rows(path)
    if not rows:
        return {"imported": 0, "skipped": 0, "errors": ["文件中没有可导入的数据"]}

    headers = list(rows[0].keys())
    header_map, dynamic_headers = build_header_map(headers)
    if "student_id" not in header_map:
        raise ValueError("导入文件必须包含学号列")
    if "name" not in header_map:
        raise ValueError("导入文件必须包含姓名或学生姓名列")

    batch_id = student_store.create_import_batch(filename or os.path.basename(path))
    imported = 0
    updated = 0
    conflicts = 0
    new_students = 0
    skipped = 0
    errors = []

    existing_dynamic = {field["field_key"] for field in student_store.get_dynamic_fields()}

    for index, row in enumerate(rows, start=2):
        fixed_data = {}
        for field, source_header in header_map.items():
            fixed_data[field] = str(row.get(source_header, "") or "").strip()

        name = fixed_data.get("name", "")
        number = fixed_data.get("student_id", "")
        if not name or not number:
            skipped += 1
            errors.append(f"第 {index} 行缺少姓名或学号，已跳过")
            continue

        try:
            existing_id = student_store.find_student_id_by_number(number)
            dynamic_values = {
                key: str(row.get(header, "") or "").strip()
                for key, header in dynamic_headers.items()
            }
            if not existing_id:
                student = student_store.save_student(fixed_data)
                for key, header in dynamic_headers.items():
                    new_value = dynamic_values.get(key, "")
                    if key in existing_dynamic and new_value:
                        student_store.set_extra_value(student["id"], key, new_value)
                    elif new_value:
                        student_store.add_import_change(
                            batch_id, student["id"], number, name, key, header, "", new_value, "new_field"
                        )
                        conflicts += 1
                imported += 1
                new_students += 1
                continue

            existing = student_store.get_student(existing_id)
            if existing.get("name") and existing.get("name") != name:
                student_store.add_import_change(
                    batch_id, existing_id, number, name, "name", "姓名",
                    existing.get("name", ""), name, "conflict"
                )
                conflicts += 1
                imported += 1
                continue

            direct_updates = {}
            for field, new_value in fixed_data.items():
                if field in ("student_id", "name") or not new_value:
                    continue
                old_value = str(existing.get(field, "") or "").strip()
                if not old_value:
                    direct_updates[field] = new_value
                elif old_value != new_value:
                    student_store.add_import_change(
                        batch_id, existing_id, number, name, field,
                        student_store.FIELD_LABELS.get(field, field),
                        old_value, new_value, "conflict"
                    )
                    conflicts += 1

            for key, header in dynamic_headers.items():
                new_value = dynamic_values.get(key, "")
                if not new_value:
                    continue
                if key not in existing_dynamic:
                    student_store.add_import_change(
                        batch_id, existing_id, number, name, key, header, "", new_value, "new_field"
                    )
                    conflicts += 1
                    continue
                old_value = existing.get("extra_values", {}).get(key, "")
                if not old_value:
                    student_store.set_extra_value(existing_id, key, new_value)
                    direct_updates[f"extra:{key}"] = new_value
                elif old_value != new_value:
                    student_store.add_import_change(
                        batch_id, existing_id, number, name, key, header, old_value, new_value, "conflict"
                    )
                    conflicts += 1

            fixed_updates = {key: value for key, value in direct_updates.items() if not key.startswith("extra:")}
            if fixed_updates:
                student_store.save_student(fixed_updates, existing_id)
            if direct_updates:
                updated += 1
            imported += 1
        except Exception as exc:
            skipped += 1
            errors.append(f"第 {index} 行导入失败: {exc}")

    student_store.update_import_batch(
        batch_id,
        imported_count=imported,
        updated_count=updated,
        conflict_count=conflicts,
        new_student_count=new_students,
        skipped_count=skipped,
    )
    return {
        "batch_id": batch_id,
        "imported": imported,
        "updated": updated,
        "new_students": new_students,
        "conflicts": conflicts,
        "skipped": skipped,
        "errors": errors,
        "recognized_columns": header_map,
        "new_field_columns": dynamic_headers,
    }
