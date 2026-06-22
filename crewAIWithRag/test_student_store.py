import os
import tempfile

import export_service
import student_store


def reset_temp_db():
    temp_dir = tempfile.mkdtemp(prefix="student_store_test_")
    student_store.DB_PATH = os.path.join(temp_dir, "students.db")
    export_service.EXPORT_DIR = os.path.join(temp_dir, "exports")
    student_store.init_db()


def run_tests():
    reset_temp_db()

    student = student_store.save_student(
        {
            "name": "张三",
            "class_name": "软件一班",
            "student_id": "20240001",
            "phone": "13800000000",
            "dorm_location": "东区",
            "dorm_room": "3-501",
            "advisor_name": "李老师",
            "gender": "男",
            "political_status": "共青团员",
            "tripartite_status": "否",
            "destination_type": "待就业",
        }
    )
    assert student["id"] > 0
    assert student["name"] == "张三"
    assert student_store.save_student({"name": "李四", "student_id": "20240003", "gender": "ХЎ"})["gender"] == "男"
    assert student_store.save_student({"name": "王五", "student_id": "20240004", "gender": "Фа"})["gender"] == "女"

    batch_result = student_store.batch_search_students(["张三", "20240003", "13800000000", "不存在", "张三"])
    assert batch_result["terms"] == ["张三", "20240003", "13800000000", "不存在"]
    assert batch_result["missing"] == ["不存在"]
    assert {item["student_id"] for item in batch_result["students"]} == {"20240001", "20240003"}

    updated = student_store.save_student(
        {"name": "张三", "class_name": "软件二班", "extra_values": {"extra_qq号": "123456"}},
        student["id"],
    )
    assert updated["class_name"] == "软件二班"
    assert updated["extra_values"]["extra_qq号"] == "123456"

    try:
        student_store.save_student({"name": "王五", "student_id": "20240001"})
        raise AssertionError("重复学号应该被拒绝")
    except ValueError:
        pass

    record = student_store.add_record(
        student["id"],
        {
            "record_type": "奖励",
            "title": "校级奖学金",
            "record_date": "2026-06-22",
            "description": "综合表现优秀",
            "source": "学生处",
        },
    )
    assert record["title"] == "校级奖学金"
    assert student_store.get_stats()["reward_count"] == 1
    listed = [item for item in student_store.list_students("20240001") if item["id"] == student["id"]][0]
    assert listed["record_count"] == 1
    assert listed["records"] == []

    activity = student_store.add_activity(
        student["id"],
        {
            "activity_date": "2026-06-22",
            "task_name": "校园志愿服务",
            "task_quantity": 2,
            "duration_hours": 3,
            "score": 6,
            "score_rule": "任务量 * 时长",
            "note": "图书馆整理资料",
        },
    )
    assert activity["score"] == 6
    assert student_store.get_activity_score(student["id"]) == 6
    assert student_store.get_stats()["activity_score"] == 6

    batch_id = student_store.create_import_batch("更新表.xlsx")
    student_store.add_import_change(
        batch_id,
        student["id"],
        "20240001",
        "张三",
        "extra_qq号",
        "QQ号",
        "123456",
        "654321",
        "conflict",
    )
    change = student_store.list_import_changes(batch_id)[0]
    student_store.apply_import_change(change["id"], "apply")
    assert student_store.get_student(student["id"])["extra_values"]["extra_qq号"] == "654321"
    assert student_store.count_import_changes(batch_id) == 0

    student_store.add_import_change(
        batch_id,
        student["id"],
        "20240001",
        "张三",
        "gender",
        "性别",
        "",
        "ХЎ",
        "conflict",
    )
    change_id = student_store.list_import_change_ids(batch_id, limit=1)[0]
    student_store.apply_import_change(change_id, "apply")
    assert student_store.get_student(student["id"])["gender"] == "男"

    student_store.batch_update_students([student["id"]], {"advisor_name": "王老师", "student_id": "不能改"})
    assert student_store.get_student(student["id"])["advisor_name"] == "王老师"
    assert student_store.get_student(student["id"])["student_id"] == "20240001"

    path = export_service.export_excel(["students", "records", "activities"], {}, ["student_id", "name", "class_name"])
    assert os.path.exists(path)
    assert path.endswith(".xlsx")
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    assert [cell.value for cell in next(wb["学生信息"].iter_rows(max_row=1))] == ["学号", "姓名", "班级"]

    student_store.delete_activity(activity["id"])
    assert student_store.get_activity_score(student["id"]) == 0

    student_store.delete_record(record["id"])
    assert len(student_store.get_student(student["id"])["records"]) == 0

    student_store.delete_student(student["id"])
    assert student_store.get_student(student["id"]) is None

    print("student_store tests passed")


if __name__ == "__main__":
    run_tests()
